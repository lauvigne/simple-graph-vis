from __future__ import annotations

import json
import re
import random
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE_URL = "https://bian.org/servicelandscape-14-0-0"
ROOT_VIEW_ID = 53866
OUTPUT_FILE = Path("bian_business_capabilities.xlsx")

NS = {"svg": "http://www.w3.org/2000/svg"}


APPLICABILITY_AXES = [
    {
        "axis": "Regulation",
        "prefix": "REG",
        "tree": [
            {
                "label": "Financial Crime Compliance",
                "children": [
                    {
                        "label": "AML",
                        "children": [
                            {"label": "Screening", "children": [{"label": "Sanctions Screening"}, {"label": "PEP Screening"}]},
                            {"label": "Monitoring", "children": [{"label": "Transaction Monitoring"}, {"label": "Alert Triage"}]},
                        ],
                    },
                    {
                        "label": "KYC",
                        "children": [
                            {"label": "Customer Due Diligence", "children": [{"label": "Initial CDD"}, {"label": "Ongoing CDD"}]},
                            {"label": "Identity Verification", "children": [{"label": "Document Check"}, {"label": "Biometric Check"}]},
                        ],
                    },
                ],
            },
            {
                "label": "Conduct & Consumer Protection",
                "children": [
                    {"label": "Suitability", "children": [{"label": "Product Suitability"}]},
                    {"label": "Disclosure", "children": [{"label": "Terms Disclosure"}, {"label": "Fees Disclosure"}]},
                    {"label": "Complaints", "children": [{"label": "Complaint Handling"}]},
                ],
            },
            {
                "label": "Prudential & Risk",
                "children": [
                    {"label": "Capital Adequacy", "children": [{"label": "Capital Reporting"}]},
                    {"label": "Liquidity", "children": [{"label": "Liquidity Reporting"}]},
                ],
            },
            {
                "label": "Data Privacy",
                "children": [
                    {"label": "GDPR", "children": [{"label": "Consent", "children": [{"label": "Consent Capture"}]}]},
                    {"label": "Data Retention", "children": [{"label": "Retention Policy"}]},
                ],
            },
        ],
    },
    {
        "axis": "Customer Segment",
        "prefix": "SEG",
        "tree": [
            {
                "label": "Retail",
                "children": [
                    {"label": "Mass Market", "children": [{"label": "Everyday Banking"}]},
                    {"label": "Affluent", "children": [{"label": "Mass Affluent"}]},
                    {"label": "Private Banking", "children": [{"label": "High Net Worth"}]},
                ],
            },
            {
                "label": "Business",
                "children": [
                    {"label": "SME", "children": [{"label": "Micro Business"}, {"label": "Small Business"}]},
                    {"label": "Mid-Corporate", "children": [{"label": "Upper Mid-Corp"}]},
                    {"label": "Large Corporate", "children": [{"label": "Multinational"}]},
                ],
            },
            {
                "label": "Institutional",
                "children": [
                    {"label": "Financial Institutions", "children": [{"label": "Banks"}, {"label": "Insurers"}]},
                    {"label": "Public Sector", "children": [{"label": "Government"}]},
                ],
            },
        ],
    },
    {
        "axis": "Product / Service",
        "prefix": "PRD",
        "tree": [
            {
                "label": "Deposits",
                "children": [
                    {"label": "Current Accounts", "children": [{"label": "Retail Current Account"}, {"label": "Business Current Account"}]},
                    {"label": "Savings", "children": [{"label": "Instant Access Savings"}, {"label": "Term Deposit"}]},
                ],
            },
            {
                "label": "Lending",
                "children": [
                    {"label": "Consumer Lending", "children": [{"label": "Personal Loan"}, {"label": "Mortgage"}]},
                    {"label": "Business Lending", "children": [{"label": "SME Lending"}, {"label": "Corporate Lending"}]},
                ],
            },
            {
                "label": "Payments",
                "children": [
                    {"label": "Domestic Payments", "children": [{"label": "Transfers"}, {"label": "Direct Debit"}]},
                    {"label": "Cross-border Payments", "children": [{"label": "SWIFT"}, {"label": "Correspondent Banking"}]},
                    {"label": "Cards", "children": [{"label": "Debit Cards"}, {"label": "Credit Cards"}]},
                ],
            },
            {
                "label": "Wealth",
                "children": [
                    {"label": "Advisory", "children": [{"label": "Financial Planning"}]},
                    {"label": "Investment Management", "children": [{"label": "Managed Portfolio"}, {"label": "Brokerage"}]},
                ],
            },
        ],
    },
    {
        "axis": "Channel",
        "prefix": "CHN",
        "tree": [
            {
                "label": "Assisted",
                "children": [
                    {"label": "Branch", "children": [{"label": "In-Branch"}]},
                    {"label": "Contact Center", "children": [{"label": "Voice"}, {"label": "Chat"}]},
                ],
            },
            {
                "label": "Digital",
                "children": [
                    {"label": "Web", "children": [{"label": "Online Banking"}]},
                    {"label": "Mobile", "children": [{"label": "Mobile App"}]},
                    {"label": "API", "children": [{"label": "Open Banking"}]},
                ],
            },
            {
                "label": "Self-Service",
                "children": [
                    {"label": "ATM", "children": [{"label": "Cash Withdrawal"}]},
                    {"label": "Kiosk", "children": [{"label": "Digital Kiosk"}]},
                ],
            },
        ],
    },
    {
        "axis": "Geography",
        "prefix": "GEO",
        "tree": [
            {
                "label": "Domestic",
                "children": [
                    {"label": "Home Market", "children": [{"label": "National"}]},
                    {"label": "Regional", "children": [{"label": "Near Domestic"}]},
                ],
            },
            {
                "label": "International",
                "children": [
                    {
                        "label": "EMEA",
                        "children": [{"label": "Western Europe"}, {"label": "Middle East and Africa"}],
                    },
                    {
                        "label": "Americas",
                        "children": [{"label": "North America"}, {"label": "Latin America"}],
                    },
                    {
                        "label": "APAC",
                        "children": [{"label": "Developed APAC"}, {"label": "Emerging APAC"}],
                    },
                ],
            },
        ],
    },
    {
        "axis": "Business Domain",
        "prefix": "DOM",
        "tree": [
            {
                "label": "Retail Banking",
                "children": [
                    {"label": "Daily Banking", "children": [{"label": "Accounts & Payments"}]},
                    {"label": "Lending", "children": [{"label": "Consumer Lending"}]},
                    {"label": "Savings & Wealth", "children": [{"label": "Investments"}]},
                ],
            },
            {
                "label": "Corporate Banking",
                "children": [
                    {"label": "Transaction Banking", "children": [{"label": "Cash Management"}, {"label": "Trade Finance"}]},
                    {"label": "Treasury", "children": [{"label": "Liquidity Management"}]},
                ],
            },
            {
                "label": "Wealth Management",
                "children": [
                    {"label": "Advisory", "children": [{"label": "Portfolio Advisory"}]},
                    {"label": "Private Banking", "children": [{"label": "High Net Worth Service"}]},
                ],
            },
        ],
    },
]


@dataclass
class Node:
    bid: str
    sem: str | None
    concept: str
    label: str
    bbox: tuple[float, float, float, float]
    children: list[str] = field(default_factory=list)
    parent: str | None = None

    @property
    def area(self) -> float:
        x1, y1, x2, y2 = self.bbox
        return max(0.0, (x2 - x1) * (y2 - y1))


def http_get(url: str) -> str:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return response.text


def load_objects_on_views() -> dict[str, list[int]]:
    text = http_get(f"{BASE_URL}/data/all_objects_on_views.js")
    match = re.search(r"var objectsOnViews = (\{.*?\});\s*var ", text, re.S)
    if not match:
        raise RuntimeError("Impossible de charger all_objects_on_views.js")
    raw = json.loads(match.group(1))
    return {str(key): value for key, value in raw.items()}


def extract_svg(html: str) -> str:
    marker = '<?xml version="1.0" encoding="UTF-8" standalone="no"?>'
    start = html.find(marker)
    if start < 0:
        raise RuntimeError("SVG introuvable dans la page BIAN")
    end = html.find("</svg>", start)
    if end < 0:
        raise RuntimeError("Balise </svg> introuvable")
    return html[start : end + len("</svg>")]


def normalize_label(text: str) -> str:
    return " ".join(text.split()).strip()


def parse_path_bbox(d: str) -> tuple[float, float, float, float]:
    tokens = re.findall(r"[A-Za-z]|-?\d+(?:\.\d+)?", d)
    i = 0
    x = y = 0.0
    min_x = min_y = float("inf")
    max_x = max_y = float("-inf")
    cmd = None

    def mark(px: float, py: float) -> None:
        nonlocal min_x, min_y, max_x, max_y
        min_x = min(min_x, px)
        min_y = min(min_y, py)
        max_x = max(max_x, px)
        max_y = max(max_y, py)

    while i < len(tokens):
        token = tokens[i]
        if re.fullmatch(r"[A-Za-z]", token):
            cmd = token
            i += 1
            continue

        if cmd in ("M", "m"):
            nx = float(tokens[i])
            ny = float(tokens[i + 1])
            i += 2
            if cmd == "m":
                x += nx
                y += ny
            else:
                x, y = nx, ny
            mark(x, y)
            continue

        if cmd in ("h", "H"):
            dx = float(token)
            x = x + dx if cmd == "h" else dx
            mark(x, y)
            i += 1
            continue

        if cmd in ("v", "V"):
            dy = float(token)
            y = y + dy if cmd == "v" else dy
            mark(x, y)
            i += 1
            continue

        if cmd in ("a", "A"):
            if i + 6 >= len(tokens):
                break
            dx = float(tokens[i + 5])
            dy = float(tokens[i + 6])
            if cmd == "a":
                x += dx
                y += dy
            else:
                x, y = dx, dy
            mark(x, y)
            i += 7
            continue

        if cmd in ("l", "L"):
            nx = float(tokens[i])
            ny = float(tokens[i + 1])
            i += 2
            if cmd == "l":
                x += nx
                y += ny
            else:
                x, y = nx, ny
            mark(x, y)
            continue

        i += 1

    if min_x == float("inf"):
        raise RuntimeError(f"Impossible de calculer la bbox pour le chemin: {d[:80]}")
    return (min_x, min_y, max_x, max_y)


def text_for_node(group: etree._Element) -> str:
    bid = group.attrib.get("bizzid")
    if not bid:
        return ""
    label_group = group.find(f"./svg:g[@bizzid='label{bid}']", NS)
    if label_group is None:
        return ""
    texts = ["".join(text.itertext()).strip() for text in label_group.findall(".//svg:text", NS)]
    return normalize_label(" ".join(part for part in texts if part))


def parse_view(view_id: int) -> tuple[list[Node], dict[str, Node]]:
    html = http_get(f"{BASE_URL}/views/view_{view_id}.html")
    svg = extract_svg(html)
    root = etree.fromstring(svg.encode("utf-8"))

    nodes: list[Node] = []
    for group in root.xpath(
        './/svg:g[@bizzconcept="CompositeGrouping" or @bizzconcept="StrategyCapability"]',
        namespaces=NS,
    ):
      bid = group.attrib.get("bizzid")
      if not bid:
          continue
      concept = group.attrib.get("bizzconcept", "")
      sem = group.attrib.get("bizzsemantic")
      label = text_for_node(group)
      path = group.find(f"./svg:g[@class='object{bid}']/svg:path", NS)
      if path is None:
          continue
      bbox = parse_path_bbox(path.attrib["d"])
      nodes.append(Node(bid=bid, sem=sem, concept=concept, label=label, bbox=bbox))

    def contains(outer: Node, inner: Node) -> bool:
        ox1, oy1, ox2, oy2 = outer.bbox
        ix1, iy1, ix2, iy2 = inner.bbox
        return ox1 <= ix1 and oy1 <= iy1 and ox2 >= ix2 and oy2 >= iy2 and outer.bid != inner.bid

    for node in nodes:
        containers = [candidate for candidate in nodes if contains(candidate, node)]
        if containers:
            parent = min(containers, key=lambda candidate: candidate.area)
            node.parent = parent.bid
            parent.children.append(node.bid)

    return nodes, {node.bid: node for node in nodes}


def direct_children(nodes: dict[str, Node], parent_bid: str) -> list[Node]:
    return sorted(
        (nodes[child_bid] for child_bid in nodes[parent_bid].children),
        key=lambda node: (node.bbox[1], node.bbox[0], node.label),
    )


def choose_detail_view(objects_on_views: dict[str, list[int]], node: Node, current_view: int) -> int | None:
    if not node.sem:
        return None
    views = [view for view in objects_on_views.get(str(node.sem), []) if view != current_view]
    return views[0] if views else None


def make_long_name(code: str, *segments: str) -> str:
    return f"{code} {' / '.join(segments)}"


def flatten_applicability_taxonomy() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    def visit(axis: str, prefix: str, node: dict[str, Any], path: list[str], code_parts: list[str], parent_code: str | None, sibling_index: int) -> None:
        current_path = path + [node["label"]]
        current_code_parts = code_parts + [str(sibling_index)]
        code = f"{prefix}-" + ".".join(current_code_parts)
        path_values = (current_path + ["", "", "", ""])[:4]
        rows.append(
            {
                "Axis": axis,
                "Level 1": path_values[0],
                "Level 2": path_values[1],
                "Level 3": path_values[2],
                "Level 4": path_values[3],
                "Code": code,
                "Long Name": f"{code} {' / '.join(current_path)}",
                "Parent Code": parent_code or "",
            }
        )
        for child_index, child in enumerate(node.get("children", []), start=1):
            visit(axis, prefix, child, current_path, current_code_parts, code, child_index)

    for axis_spec in APPLICABILITY_AXES:
        axis = axis_spec["axis"]
        prefix = axis_spec["prefix"]
        for root_index, root_node in enumerate(axis_spec["tree"], start=1):
            visit(axis, prefix, root_node, [], [], None, root_index)

    return rows


def build_capability_catalog(rows: list[dict[str, str]]) -> dict[str, Any]:
    l1_to_l2: dict[str, list[str]] = {}
    l2_to_l3: dict[str, list[str]] = {}
    l2_to_l1: dict[str, str] = {}
    l2_long_by_label: dict[str, str] = {}
    l3_long_by_label: dict[str, str] = {}

    for row in rows:
        l1 = row["Business Capability (L1)"].strip()
        l2 = row["Business Capability (L2)"].strip()
        l2_long = row["Business Capability (L2) long name"].strip()
        l3 = row["Business Capability (L3)"].strip()
        l3_long = row["Business Capability (L3) long name"].strip()

        if l1 and l2:
            l1_to_l2.setdefault(l1, [])
            if l2 not in l1_to_l2[l1]:
                l1_to_l2[l1].append(l2)
            l2_to_l1.setdefault(l2, l1)
            if l2_long:
                l2_long_by_label.setdefault(l2, l2_long)

        if l2 and l3:
            l2_to_l3.setdefault(l2, [])
            if l3 not in l2_to_l3[l2]:
                l2_to_l3[l2].append(l3)
            if l3_long:
                l3_long_by_label.setdefault(l3, l3_long)

    for values in l1_to_l2.values():
        values.sort()
    for values in l2_to_l3.values():
        values.sort()

    return {
        "l1_to_l2": l1_to_l2,
        "l2_to_l3": l2_to_l3,
        "l2_to_l1": l2_to_l1,
        "l2_long_by_label": l2_long_by_label,
        "l3_long_by_label": l3_long_by_label,
        "all_l1": sorted(l1_to_l2),
        "all_l2": sorted(l2_to_l1),
        "eligible_l2": sorted([l2 for l2, l3s in l2_to_l3.items() if l3s]),
        "eligible_l3": sorted(l3_long_by_label),
    }


def make_app_name(entity: str, family: str, index: int) -> tuple[str, str]:
    code = f"{entity}-APP-{index:04d}"
    display = f"{family} {index:04d}"
    return code, display


def choose_area(rng: random.Random, catalog: dict[str, Any], theme_keywords: list[str]) -> str:
    matches = [
        area
        for area in catalog["all_l1"]
        if any(keyword in area.lower() for keyword in theme_keywords)
    ]
    pool = matches or catalog["all_l1"]
    return rng.choice(pool)


def choose_l2_labels(rng: random.Random, catalog: dict[str, Any], area: str, count: int) -> list[str]:
    candidates = catalog["l1_to_l2"].get(area, catalog["eligible_l2"])
    if not candidates:
        return []
    count = max(1, min(count, len(candidates)))
    if count == len(candidates):
        return candidates[:]
    return rng.sample(candidates, count)


def choose_l3_labels(rng: random.Random, catalog: dict[str, Any], l2_labels: list[str], count: int) -> list[str]:
    pool: list[str] = []
    for l2 in l2_labels:
        for l3 in catalog["l2_to_l3"].get(l2, []):
            if l3 not in pool:
                pool.append(l3)
    if not pool:
        return []
    count = max(1, min(count, len(pool)))
    if count == len(pool):
        return pool[:]
    return rng.sample(pool, count)


def generate_bcm_rows(
    entity: str,
    target_count: int,
    catalog: dict[str, Any],
) -> list[dict[str, str]]:
    family_specs = {
        "E1": [
            ("Retail Payments Hub", ["payment", "card", "channel", "customer"]),
            ("Digital Banking Platform", ["customer", "channel", "service", "account"]),
            ("Cards and Wallets Engine", ["card", "payment", "channel"]),
            ("Customer Onboarding Suite", ["customer", "service", "channel"]),
            ("Branch Service Workbench", ["service", "customer", "channel"]),
            ("Retail Lending Origination", ["loan", "customer", "service"]),
            ("Collections and Recoveries", ["loan", "risk", "service"]),
            ("Fraud Monitoring Console", ["risk", "security", "payment"]),
            ("Savings Product Hub", ["saving", "deposit", "customer"]),
            ("Notifications and Statements", ["service", "customer", "channel"]),
            ("Open Banking Gateway", ["api", "channel", "customer"]),
            ("Customer Care CRM", ["customer", "service"]),
        ],
        "E2": [
            ("Corporate Cash Management", ["cash", "corporate", "treasury"]),
            ("Trade Finance Workbench", ["trade", "corporate", "risk"]),
            ("Treasury Dealing Desk", ["treasury", "liquidity", "risk"]),
            ("Commercial Lending Factory", ["lending", "corporate", "risk"]),
            ("Supplier Finance Hub", ["corporate", "lending", "payment"]),
            ("FX and Liquidity Platform", ["liquidity", "treasury", "payment"]),
            ("Corporate Onboarding Portal", ["corporate", "customer", "service"]),
            ("Collections Factory", ["corporate", "risk", "service"]),
            ("Guarantees Engine", ["corporate", "risk", "service"]),
            ("Payments Factory", ["payment", "corporate", "channel"]),
            ("Corporate CRM", ["corporate", "customer", "service"]),
            ("Client Profitability Studio", ["corporate", "data", "service"]),
        ],
        "E3": [
            ("Wealth Advisory Hub", ["wealth", "advisory", "customer"]),
            ("Portfolio Management Suite", ["wealth", "investment", "risk"]),
            ("Private Banking CRM", ["private", "wealth", "customer"]),
            ("Investment Products Factory", ["wealth", "product", "service"]),
            ("Suitability and Risk Engine", ["wealth", "risk", "regulation"]),
            ("Research Distribution Portal", ["wealth", "channel", "customer"]),
            ("Client Reporting Hub", ["wealth", "report", "customer"]),
            ("Mandate Management Workbench", ["wealth", "investment", "service"]),
            ("Wealth Onboarding Flow", ["wealth", "customer", "service"]),
            ("Fee Billing and Performance", ["wealth", "finance", "service"]),
            ("Alternative Investments Desk", ["wealth", "investment"]),
            ("Estate Planning Studio", ["wealth", "customer", "service"]),
        ],
        "E4": [
            ("Enterprise Service Hub", ["service", "customer", "channel"]),
            ("Product Factory", ["product", "service", "data"]),
            ("Process Orchestration Tower", ["service", "risk", "data"]),
            ("Operations Control Tower", ["service", "risk", "report"]),
            ("Shared Payments Platform", ["payment", "channel", "service"]),
            ("Master Data Hub", ["data", "service", "customer"]),
            ("Risk and Compliance Hub", ["risk", "regulation", "service"]),
            ("Enterprise Reporting Suite", ["report", "data", "service"]),
            ("Identity and Access Platform", ["security", "regulation", "customer"]),
            ("Workflow Engine", ["service", "data", "channel"]),
            ("Customer 360", ["customer", "data", "service"]),
            ("Integration Platform", ["api", "data", "service"]),
        ],
    }

    families = family_specs[entity]
    rng = random.Random(f"{entity}:{target_count}")
    anchor_rng = random.Random(f"{entity}:capability-anchors")
    family_anchors: dict[str, list[str]] = {}
    for family, keywords in families:
        area = choose_area(anchor_rng, catalog, keywords)
        anchors = choose_l2_labels(anchor_rng, catalog, area, 4)
        if not anchors:
            anchors = anchor_rng.sample(catalog["eligible_l2"], k=min(4, len(catalog["eligible_l2"])))
        family_anchors[family] = anchors

    rows: list[dict[str, str]] = []

    for index in range(1, target_count + 1):
        family, keywords = families[(index - 1) % len(families)]
        code, display = make_app_name(entity, family, index)
        mapping_kind = "l2" if rng.random() < 0.4 else "l3"
        anchor_l2 = family_anchors.get(family, [])

        if mapping_kind == "l2":
            l2_count = 1 if rng.random() < 0.9 else 2
            l2_labels = rng.sample(anchor_l2, k=min(l2_count, len(anchor_l2))) if anchor_l2 else []
            if not l2_labels:
                l2_labels = [rng.choice(catalog["all_l2"])]
            l2_values = [catalog["l2_long_by_label"].get(label, label) for label in l2_labels]
            l3_values: list[str] = []
        else:
            l2_count = 1 if rng.random() < 0.75 else 2
            l2_labels = rng.sample(anchor_l2, k=min(l2_count, len(anchor_l2))) if anchor_l2 else []
            if not l2_labels:
                l2_labels = [rng.choice(catalog["eligible_l2"])]
            l3_target = rng.randint(1, 4)
            l3_labels = choose_l3_labels(rng, catalog, l2_labels, l3_target)
            if not l3_labels and catalog["eligible_l3"]:
                l3_labels = rng.sample(catalog["eligible_l3"], k=min(l3_target, len(catalog["eligible_l3"])))
            l2_values = []
            l3_values = [catalog["l3_long_by_label"].get(label, label) for label in l3_labels]

        rows.append(
            {
                "Application Code": code,
                "Application Name": f"{family} {entity[-1]}",
                "Application Display Name": display,
                "BIAN L2": ", ".join(l2_values),
                "BIAN L3": ", ".join(l3_values),
            }
        )

    return rows


def write_bcm_sheet(wb: Workbook, entity: str, target_count: int, catalog: dict[str, Any]) -> int:
    ws = wb.create_sheet(f"{entity}-BCM")
    headers = [
        "Application Code",
        "Application Name",
        "Application Display Name",
        "BIAN L2",
        "BIAN L3",
    ]

    header_fill = PatternFill("solid", fgColor="14532D")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(vertical="top", wrap_text=True)

    rows = generate_bcm_rows(entity, target_count, catalog)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment

    widths = {
        "A": 20,
        "B": 34,
        "C": 34,
        "D": 64,
        "E": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    return len(rows)


def write_taxonomy_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Applicability Taxonomy")
    headers = [
        "Axis",
        "Level 1",
        "Level 2",
        "Level 3",
        "Level 4",
        "Code",
        "Long Name",
        "Parent Code",
    ]
    rows = flatten_applicability_taxonomy()

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(vertical="top", wrap_text=True)

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment

    widths = {
        "A": 24,
        "B": 28,
        "C": 30,
        "D": 30,
        "E": 30,
        "F": 16,
        "G": 80,
        "H": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_rows() -> list[dict[str, str]]:
    objects_on_views = load_objects_on_views()
    root_nodes, root_map = parse_view(ROOT_VIEW_ID)
    root_areas = [
        node
        for node in root_nodes
        if node.concept == "CompositeGrouping" and node.parent is None
    ]
    root_areas = sorted(root_areas, key=lambda node: (node.bbox[0], node.bbox[1], node.label))

    rows: list[dict[str, str]] = []

    for area_index, area in enumerate(root_areas, start=1):
        area_code = str(area_index)
        l2_candidates = [
            root_map[child_bid]
            for child_bid in area.children
            if root_map[child_bid].concept == "StrategyCapability"
        ]
        l2_candidates = sorted(l2_candidates, key=lambda node: (node.bbox[1], node.bbox[0], node.label))

        for l2_index, l2_node in enumerate(l2_candidates, start=1):
            l2_code = f"{area_code}.{l2_index}"
            l2_long = make_long_name(l2_code, area.label, l2_node.label)
            detail_view = choose_detail_view(objects_on_views, l2_node, ROOT_VIEW_ID)

            l3_nodes: list[Node] = []
            if detail_view is not None:
                detail_nodes, detail_map = parse_view(detail_view)
                detail_roots = [
                    node
                    for node in detail_nodes
                    if node.concept == "StrategyCapability" and node.parent is None
                ]
                if detail_roots:
                    root = sorted(detail_roots, key=lambda node: (node.bbox[0], node.bbox[1], node.label))[0]
                    l3_nodes = sorted(
                        (detail_map[child_bid] for child_bid in root.children if detail_map[child_bid].concept == "StrategyCapability"),
                        key=lambda node: (node.bbox[1], node.bbox[0], node.label),
                    )

            if not l3_nodes:
                rows.append(
                    {
                        "Business Capability (L1)": area.label,
                        "Business Capability (L2)": l2_node.label,
                        "Business Capability (L2) long name": l2_long,
                        "Business Capability (L3)": "",
                        "Business Capability (L3) long name": "",
                    }
                )
                continue

            for l3_index, l3_node in enumerate(l3_nodes, start=1):
                l3_code = f"{l2_code}.{l3_index}"
                rows.append(
                    {
                        "Business Capability (L1)": area.label,
                        "Business Capability (L2)": l2_node.label,
                        "Business Capability (L2) long name": l2_long,
                        "Business Capability (L3)": l3_node.label,
                        "Business Capability (L3) long name": make_long_name(l3_code, area.label, l2_node.label, l3_node.label),
                    }
                )

    return rows


def write_workbook(rows: list[dict[str, str]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "BIAN Capabilities"
    catalog = build_capability_catalog(rows)

    headers = [
        "Business Capability (L1)",
        "Business Capability (L2)",
        "Business Capability (L2) long name",
        "Business Capability (L3)",
        "Business Capability (L3) long name",
    ]

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    thin_alignment = Alignment(vertical="top", wrap_text=True)

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = thin_alignment

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = thin_alignment

    widths = {
        "A": 34,
        "B": 34,
        "C": 62,
        "D": 36,
        "E": 76,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    write_taxonomy_sheet(wb)

    bcm_counts = {
        "E1-BCM": write_bcm_sheet(wb, "E1", 500, catalog),
        "E2-BCM": write_bcm_sheet(wb, "E2", 800, catalog),
        "E3-BCM": write_bcm_sheet(wb, "E3", 1200, catalog),
        "E4-BCM": write_bcm_sheet(wb, "E4", 1500, catalog),
    }

    meta = wb.create_sheet("Metadata")
    meta["A1"] = "Source"
    meta["B1"] = f"{BASE_URL}/views/view_{ROOT_VIEW_ID}.html"
    meta["A2"] = "Generated at"
    meta["B2"] = __import__("datetime").datetime.utcnow().isoformat() + "Z"
    meta["A3"] = "Rows"
    meta["B3"] = len(rows)
    meta["A4"] = "Applicability rows"
    meta["B4"] = len(flatten_applicability_taxonomy())
    meta["A5"] = "BCM rows"
    meta["B5"] = json.dumps(bcm_counts, ensure_ascii=False)
    meta["A6"] = "Notes"
    meta["B6"] = "Hierarchy extracted from BIAN InSite top-level and detailed capability views, plus a synthetic multi-axis applicability taxonomy and synthetic entity BCM mappings."
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 90

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def main() -> None:
    rows = build_rows()
    output = write_workbook(rows)
    print(f"Wrote {output} with {len(rows)} rows")


if __name__ == "__main__":
    main()
